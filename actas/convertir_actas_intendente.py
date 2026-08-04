import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Cambia estas rutas según tu PC
CARPETA_ENTRADA = Path(r"C:\actas\csv")
CARPETA_SALIDA = Path(r"C:\actas\excel")

FILAS_SALIDA = [
    "LISTA 2A",
    "LISTA 2M",
    "LISTA 5",
    "LISTA 7",
    "NULOS",
    "BLANCO",
    "A COMPUTAR",
]


def extraer_ultimo_numero(nombre_archivo: str) -> str:
    """
    Ejemplo:
    261.11.15.427.1.1-INT-acta-recuento.csv -> 1.xlsx
    Toma el último número antes del primer guion.
    """
    stem = Path(nombre_archivo).stem
    antes_guion = stem.split("-")[0]
    numeros = re.findall(r"\d+", antes_guion)
    return numeros[-1] if numeros else stem


def convertir_numero(valor):
    if valor is None:
        return 0

    valor = str(valor).strip().replace('.', '').replace(',', '.')

    if valor == '':
        return 0

    try:
        numero = float(valor)
        return int(numero) if numero.is_integer() else numero
    except ValueError:
        return 0


def normalizar(texto: str) -> str:
    texto = str(texto or '').strip().upper()
    cambios = str.maketrans('ÁÉÍÓÚÑ', 'AEIOUN')
    return texto.translate(cambios)


def leer_csv_acta_intendente(ruta_csv: Path) -> dict:
    datos = {fila: 0 for fila in FILAS_SALIDA}

    with ruta_csv.open('r', encoding='utf-8-sig', newline='') as archivo:
        lector = csv.reader(archivo)
        next(lector, None)  # encabezado

        for fila in lector:
            if not fila:
                continue

            primera_col = normalizar(fila[0] if len(fila) > 0 else '')
            clase = normalizar(fila[1] if len(fila) > 1 else '')
            nro = normalizar(fila[2] if len(fila) > 2 else '')
            agrupacion = normalizar(fila[3] if len(fila) > 3 else '')

            # Listas normales: el voto de intendente viene en la última columna INT
            if clase == 'LISTA':
                nombre_lista = f"LISTA {nro}"

                if nombre_lista in datos:
                    datos[nombre_lista] = convertir_numero(fila[4] if len(fila) > 4 else 0)
                continue

            # Blanco: en el CSV viene como fila Blanco y el total queda en la última columna disponible
            if clase == 'BLANCO' or nro == 'VOTO EN BLANCO' or agrupacion == 'VOTO EN BLANCO':
                datos['BLANCO'] = convertir_numero(fila[-1] if len(fila) > 0 else 0)
                continue

            # Nulos
            if primera_col == 'VOTOS NULOS':
                datos['NULOS'] = convertir_numero(fila[1] if len(fila) > 1 else 0)
                continue

            # A computar
            if primera_col == 'VOTOS A COMPUTAR':
                datos['A COMPUTAR'] = convertir_numero(fila[1] if len(fila) > 1 else 0)
                continue

    return datos


def crear_excel_intendente(datos: dict, ruta_salida: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Worksheet'

    ws.append(['lista', 'voto'])

    for nombre_fila in FILAS_SALIDA:
        ws.append([nombre_fila, datos.get(nombre_fila, 0)])

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='808080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)


def convertir_archivo(ruta_csv: Path, carpeta_salida: Path):
    numero = extraer_ultimo_numero(ruta_csv.name)
    ruta_salida = carpeta_salida / f'{numero}.xlsx'

    datos = leer_csv_acta_intendente(ruta_csv)
    crear_excel_intendente(datos, ruta_salida)

    print(f'OK: {ruta_csv.name} -> {ruta_salida.name}')


def convertir_carpeta(carpeta_entrada: Path, carpeta_salida: Path):
    archivos = sorted(carpeta_entrada.glob('*.csv'))

    if not archivos:
        print(f'No se encontraron CSV en: {carpeta_entrada}')
        return

    for ruta_csv in archivos:
        convertir_archivo(ruta_csv, carpeta_salida)


if __name__ == '__main__':
    convertir_carpeta(CARPETA_ENTRADA, CARPETA_SALIDA)
