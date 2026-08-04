import csv
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Cambia estas rutas según tu PC
CARPETA_ENTRADA = Path(r"C:\actas\csv")
CARPETA_SALIDA = Path(r"C:\actas\excel")

FILAS_SALIDA = [
    "LISTA 2A",
    "LISTA 2M",
    "LISTA 5",
    "LISTA 7",
    "NULO",
    "BLANCO",
    "A COMPUTAR",
]


def extraer_ultimo_numero(nombre_archivo: str) -> str:
    """
    Ejemplo:
    261.11.15.196.1.28-JUN-acta-recuento.csv  -> 28
    Toma el último número antes de -JUN-acta-recuento.
    """
    stem = Path(nombre_archivo).stem
    antes_guion = stem.split("-")[0]
    numeros = re.findall(r"\d+", antes_guion)
    return numeros[-1] if numeros else stem


def convertir_numero(valor):
    if valor is None:
        return 0
    valor = str(valor).strip().replace('.', '').replace(',', '.')
    if valor == '':
        return 0
    try:
        numero = float(valor)
        return int(numero) if numero.is_integer() else numero
    except ValueError:
        return 0


def leer_csv_acta(ruta_csv: Path) -> dict:
    datos = {fila: [0] * 12 for fila in FILAS_SALIDA}

    with ruta_csv.open('r', encoding='utf-8-sig', newline='') as archivo:
        lector = csv.reader(archivo)
        next(lector, None)  # encabezado

        for fila in lector:
            if not fila:
                continue

            primera_col = fila[0].strip().upper() if len(fila) > 0 else ''
            clase = fila[1].strip().upper() if len(fila) > 1 else ''
            nombre = fila[2].strip().upper() if len(fila) > 2 else ''

            # Listas normales: columnas 1 al 12 están desde la posición 3 hasta 14
            if clase == 'LISTA' and nombre in datos:
                votos = fila[3:15]
                datos[nombre] = [convertir_numero(x) for x in votos]
                continue

            # Voto en blanco: en tu CSV trae solo el total. Lo colocamos en columna 1.
            if clase == 'BLANCO' or nombre == 'VOTO EN BLANCO':
                datos['BLANCO'][0] = convertir_numero(fila[3] if len(fila) > 3 else 0)
                continue

            # Votos nulos
            if primera_col == 'VOTOS NULOS':
                datos['NULO'][0] = convertir_numero(fila[1] if len(fila) > 1 else 0)
                continue

            # Votos a computar
            if primera_col == 'VOTOS A COMPUTAR':
                datos['A COMPUTAR'][0] = convertir_numero(fila[1] if len(fila) > 1 else 0)
                continue

    return datos


def crear_excel(datos: dict, ruta_salida: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja1'

    encabezado = ['lista'] + [str(i) for i in range(1, 13)]
    ws.append(encabezado)

    for nombre_fila in FILAS_SALIDA:
        ws.append([nombre_fila] + datos.get(nombre_fila, [0] * 12))

    # Formato básico parecido al modelo
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='808080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 18
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 8

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)


def convertir_archivo(ruta_csv: Path, carpeta_salida: Path):
    numero = extraer_ultimo_numero(ruta_csv.name)
    ruta_salida = carpeta_salida / f'{numero}.xlsx'

    datos = leer_csv_acta(ruta_csv)
    crear_excel(datos, ruta_salida)

    print(f'OK: {ruta_csv.name} -> {ruta_salida.name}')


def convertir_carpeta(carpeta_entrada: Path, carpeta_salida: Path):
    archivos = sorted(carpeta_entrada.glob('*.csv'))

    if not archivos:
        print(f'No se encontraron CSV en: {carpeta_entrada}')
        return

    for ruta_csv in archivos:
        convertir_archivo(ruta_csv, carpeta_salida)


if __name__ == '__main__':
    convertir_carpeta(CARPETA_ENTRADA, CARPETA_SALIDA)
